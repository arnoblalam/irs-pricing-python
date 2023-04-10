#!/usr/bin/env python3

import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def format_excel_table(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Apply header formatting
    header_font = Font(bold=True, size=12)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply number formatting for numeric cells
    num_format = '#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.number_format = num_format
            cell.alignment = Alignment(horizontal='right', vertical='center')

    # Apply text formatting for text cells
    text_font = Font(size=11)
    for cell in ws['A']:
        cell.font = text_font

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if cell.row == 1:
                cell_width = len(str(cell.value)) * 1.25
            else:
                cell_width = len(str(cell.value)) * 1.1
            if cell_width > max_length:
                max_length = cell_width
        ws.column_dimensions[column].width = max_length

    wb.save(file_path)

input_files = {
    "USD": "data/raw/trades/USD/USD_20130311-20130324.xlsx",
    "CAD": "data/raw/trades/CAD/CAD_20130311-20130324.xlsx",
}

output_file = "results/tables/tab3_phase_1a_descriptive.xlsx"


def calculate_statistics(file_path, currency):
    # Read the Excel file
    df = pd.read_excel(file_path)

    # Filter out any transactions in the Type column that are of the type 'IRS Fix-Fix'
    df = df[df["Type"] != "IRS Fix-Fix"]

    # Filter the DataFrame by the given currency
    df = df[df["Curr"] == currency]

    # Calculate the notional value for different floating leg references
    df["Floating Leg"] = df.apply(lambda row: row["Leg 1"] if row["Leg 1"] != "FIXED" else row["Leg 2"], axis=1)

    # Combine all floating legs containing the word LIBOR
    df["Floating Leg"] = df["Floating Leg"].apply(lambda x: "LIBOR" if isinstance(x, str) and "LIBOR" in x else x)

    # Calculate the total number and total notional value of cleared and uncleared swaps
    cleared_stats = df[df["Clr"] == "C"].groupby("Floating Leg").agg({"Not.": "sum", "Clr": "count"})
    uncleared_stats = df[df["Clr"] == "U"].groupby("Floating Leg").agg({"Not.": "sum", "Clr": "count"})

    # Calculate the notional value for different floating leg references
    notional_by_floating_leg = df.groupby("Floating Leg")["Not."].sum()

    # Combine the statistics into a single DataFrame
    stats_df = notional_by_floating_leg.reset_index().rename(columns={"Not.": "Notional Value"})

    stats_df["Currency"] = currency
    stats_df["Cleared"] = stats_df["Floating Leg"].apply(lambda x: cleared_stats.loc[x, "Not."] if x in cleared_stats.index else 0)
    stats_df["Uncleared"] = stats_df["Floating Leg"].apply(lambda x: uncleared_stats.loc[x, "Not."] if x in uncleared_stats.index else 0)
    stats_df["Cleared Count"] = stats_df["Floating Leg"].apply(lambda x: cleared_stats.loc[x, "Clr"] if x in cleared_stats.index else 0)
    stats_df["Uncleared Count"] = stats_df["Floating Leg"].apply(lambda x: uncleared_stats.loc[x, "Clr"] if x in uncleared_stats.index else 0)

    return stats_df

def create_excel_table(usd_stats, cad_stats):
    # Combine the USD and CAD statistics into a single DataFrame
    combined_stats = pd.concat([usd_stats, cad_stats], ignore_index=True)

    # Reorder columns and convert notional values to billions
    combined_stats = combined_stats[['Currency', 'Floating Leg', 'Cleared Count', 'Cleared', 'Uncleared Count', 'Uncleared']]
    combined_stats[['Cleared', 'Uncleared']] /= 1e6

    # Format the numbers
    format_dict = {
        "Cleared Count": "{:,.0f}",
        "Cleared": "${:,.2f}",
        "Uncleared Count": "{:,.0f}",
        "Uncleared": "${:,.2f}",
    }

    combined_stats = combined_stats.style.format(format_dict)

    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(output_file) as writer:
        combined_stats.to_excel(writer, index=False, sheet_name="Summary")

# Calculate statistics for USD and CAD
usd_stats = calculate_statistics(input_files["USD"], "USD")
cad_stats = calculate_statistics(input_files["CAD"], "CAD")

# Create the Excel table and save it
create_excel_table(usd_stats, cad_stats)
format_excel_table(output_file)
